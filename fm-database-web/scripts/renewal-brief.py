#!/usr/bin/env python3
"""Everything needed to write one client's renewal letter — and nothing invented.

Deterministic and free: files and dates only, never a model call. The letter
itself is authored in chat (see .claude/skills/author-renewal), so this is the
half that must not be improvised. Anything absent here is absent from the
letter.
"""
import datetime as dt
import glob
import json
import pathlib
import re
import sys

import yaml

ROOT = pathlib.Path.home() / "fm-plans"
TODAY = dt.date.today()


def as_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str):
        try:
            return dt.date.fromisoformat(v[:10])
        except ValueError:
            return None
    return None


def person(cid):
    for base in ("clients", "prospects"):
        p = ROOT / base / cid / "client.yaml"
        if p.exists():
            return yaml.safe_load(p.read_text()) or {}, p.parent
    return {}, None


def nbhwc_calls(name):
    """Logged coaching calls — the topics are the coach's own words."""
    try:
        import openpyxl
    except ImportError:
        return []
    f = pathlib.Path.home() / "NBHWC Certification/NBHWC-Coaching-Log-Shivani-Hariharan.xlsx"
    if not f.exists():
        return []
    wb = openpyxl.load_workbook(f, data_only=True)
    code = None
    for r in wb["KEY - DO NOT SUBMIT"].iter_rows(min_row=2, values_only=True):
        if any(name.split()[0].lower() in str(x).lower() for x in r if x):
            code = next((str(x) for x in r if x and str(x).startswith("C")), None)
    if not code:
        return []
    out = []
    for sheet in ("Coaching Log", "Pre-PSA (not loggable)"):
        for r in wb[sheet].iter_rows(min_row=2, values_only=True):
            if any(str(x).strip() == code for x in r if x):
                vals = [str(x) for x in r if x]
                topic = max(vals, key=len)
                date = next((v for v in vals if re.match(r"\d{2}/\d{2}/\d{4}", v)), "")
                out.append({"date": date, "topic": topic})
    return out[-4:]


def main(cid):
    c, cdir = person(cid)
    if not c:
        print(json.dumps({"error": f"no such client: {cid}"}))
        return 1
    name = c.get("display_name") or cid

    plan, slug, ends = None, None, None
    for f in sorted(glob.glob(str(ROOT / "published" / "*.yaml"))):
        d = yaml.safe_load(open(f)) or {}
        if d.get("client_id") != cid:
            continue
        weeks = d.get("plan_period_weeks")
        eff = as_date(d.get("meal_plan_started_on")) or (
            as_date(d.get("plan_period_start")) + dt.timedelta(days=3)
            if as_date(d.get("plan_period_start"))
            else None
        )
        if not (weeks and eff):
            continue
        end = eff + dt.timedelta(weeks=weeks)
        if ends is None or end > ends:
            plan, ends, slug = d, end, re.sub(r"-v\d+\.yaml$", "", pathlib.Path(f).name)

    msq, adherence = [], []
    for f in sorted(glob.glob(str(cdir / "sessions" / "*.yaml"))) if cdir else []:
        d = yaml.safe_load(open(f)) or {}
        pc = str(d.get("presenting_complaints") or "")
        if "msq" in pc.lower():
            m = re.search(r"total (\d+).*?Highest categories: ([^\n]+)", pc, re.S)
            w = re.search(r"\[week: (\d+)\]", pc)
            if m:
                msq.append({"week": int(w.group(1)) if w else None,
                            "total": int(m.group(1)), "top": m.group(2).strip()})
        if "adherence" in pc.lower() or "poll" in pc.lower():
            adherence.append({"date": str(d.get("date")),
                              "note": (str(d.get("coach_notes") or pc))[:220]})

    labs = []
    for lo in (plan or {}).get("lab_orders") or []:
        reason = str(lo.get("reason") or "")
        due = re.search(r"[Rr]etest at (\d+) weeks", reason)
        labs.append({"reason": reason[:200],
                     "retest_weeks": int(due.group(1)) if due else None})

    # Weight series. Present so the letter can speak about the scale — and so
    # the gate can refuse a kg figure when there is nothing to support it.
    weights = []
    for s in (c.get("health_snapshots") or []):
        m = (s.get("measurements") or {})
        if m.get("weight_kg") is not None:
            weights.append({"date": str(s.get("date")), "weight_kg": m["weight_kg"]})

    household = []
    surname = name.strip().split()[-1].lower() if " " in name else ""
    if surname:
        for f in glob.glob(str(ROOT / "clients/*/client.yaml")):
            o = yaml.safe_load(open(f)) or {}
            n = o.get("display_name") or ""
            if n and n != name and n.strip().split()[-1].lower() == surname:
                household.append(n)

    print(json.dumps({
        "client_id": cid, "name": name,
        "email": c.get("email"), "diet": c.get("dietary_preference"),
        "conditions": c.get("active_conditions") or [],
        "goals": c.get("goals") or [],
        "plan": {"slug": slug, "weeks": (plan or {}).get("plan_period_weeks"),
                 "ends_on": str(ends) if ends else None,
                 "days_left": (ends - TODAY).days if ends else None},
        "msq": msq, "adherence": adherence[-4:], "weights": weights,
        "labs_ordered": labs, "calls": nbhwc_calls(name),
        "household_also_renewing": household,
        "supplements": len((plan or {}).get("supplement_protocol") or []),
        "practices": len((plan or {}).get("lifestyle_practices") or []),
        "NOTE": "Every figure in the letter must appear above. Prices are NOT here — ask the coach.",
    }, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1] if len(sys.argv) > 1 else ""))
